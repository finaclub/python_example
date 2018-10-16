
import os 
f = open('test.txt', 'w')                # Write mode file open
f.write("This is a test")                # Write
f.save()                                 # save
f.close()                                # close


#Fuction
def myfn(x):
    if x > 100:
        print("over 100")
    else:
        print("under 100")
        
myfn(120)
myfn(90)

#Method
input = "This is a sample"
result = input.upper()                # upperはmethod
print(result)

# statistics
import statistics
 
data = [3, 7, 4, 6]
print (statistics.mean(data))

# openpyxl
from openpyxl import Workbook
import datetime

wb = Workbook()                       # workbook 作成 
ws = wb.active                        # workSheet 選択：最初のシートが選択される
ws['A1'] = 42
ws.append([1, 2, 3])                  # A2, B2, C2 に　１，２，３、が入る
ws['A3'] = datetime.datetime.now()    # A3 セルに現在の時刻
wb.save("sample.xlsx")                # ファイルネームを入れて保存
wb.close()
